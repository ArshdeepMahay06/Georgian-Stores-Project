import openpyxl, string, time
import shutil

class ExcelProcessor:

    def create_excel_file(self, keywords, source_file):

        # XYZ this is the Source file
        workbook = openpyxl.load_workbook(source_file, data_only=True)
        sheet = None

        # loading the required sheet
        if('GSProposedCourse MaterialList' in workbook.sheetnames):
            sheetIndex = workbook.sheetnames.index('GSProposedCourse MaterialList')
            sheet = workbook[workbook.sheetnames[sheetIndex]]
        else:
            sheet = workbook.active

        max_row = sheet.max_row
        max_column = sheet.max_column

        columnHeadings = []

        letters = string.ascii_uppercase

        for i in range(0,max_column):
            columnHeadings.append(sheet[letters[i]+'1'].value)

        try:
            DescriptionColumn = letters[columnHeadings.index('Description')]
        except ValueError:
            return -1
        
        inputKeywords = keywords
        inputKeywordsList = inputKeywords.lower().split(',')


        # Now the main Game !! ITERATING THE WHOLE EXCEL FILE
        favourableRows = []
        totalRecordsFoundCount = 0

        for i in range(2,max_row):
            cellValue = sheet[DescriptionColumn+str(i)].value.lower()
            for j in inputKeywordsList:
                if(j == ''):
                    continue
                if(j in cellValue):
                    favourableRows.append(i)
                    totalRecordsFoundCount += 1
                    break ;

        
        wasFileAvailable = False
        try:
            outputWorkbook = openpyxl.load_workbook(r'..\file_management\outputFile.xlsx')
            wasFileAvailable = True
            sheet = outputWorkbook.active
            
        except FileNotFoundError:
            outputWorkbook = openpyxl.Workbook()
            sheet = outputWorkbook.active
            sheet['A1'] = "Date"
            sheet['B1'] = "Input"
            sheet['C1'] = "Total Records Found"
        except PermissionError:
            print("Please close the Excel file first")
        

        
        # This is the output workbook
        

        availableRow = str(sheet.max_row + 1)

        timeNow = ','.join(time.ctime(time.time()).split()[1:])

        sheet['A'+availableRow] = timeNow
        sheet['B'+availableRow] = keywords
        sheet['C'+availableRow] = totalRecordsFoundCount

        
        
        if not wasFileAvailable:
            outputWorkbook.save(r'..\file_management\outputFile.xlsx')
        else:
            outputWorkbook.save(r'..\file_management\outputFile.xlsx')
            outputWorkbook.close()

        return totalRecordsFoundCount
